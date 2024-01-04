import time

import numpy as np
import pandas as pd
import time
import torch
import os
import math
import random 
import torch.nn as nn
import torch.nn.functional as F
import torch.distributions as dist 
from torch.optim import Adam, Adagrad, NAdam, SGD, RMSprop
from tensorboardX import SummaryWriter
from normalization import  Normalization, RewardScaling
from sklearn.preprocessing import RobustScaler
from openpyxl import load_workbook

class PPO:

	def __init__(self, policy_class, critic_class, env, tensorboard_write, excel_sheet, **hyperparameters):
		"""
			policy_class is network of actor and critic
		"""
	
		# Initialize hyperparameters for training with PPO
		self._init_hyperparameters(hyperparameters)
		torch.set_printoptions(profile="full")
		

		# GPU 
		self.device = torch.device("cuda:2" if torch.cuda.is_available() else "cpu")
	
		# Extract environment information 
		self.env = env
		self.obs_dim = 7                               # depend on s
		self.batch_obs_dim = 15
		self.act_dim = 1                                # controller ut 

		actor_configs = {
			"seq_len" : 15,
			"pred_len" : 1,
			"kernel_size" : 3,
			"individual": True,
			"enc_in" : 7
		}

		# critic_configs = {
		# 	"seq_len" : 10,
		# 	"pred_len" : 1,
		# 	"kernel_size" : 25,
		# 	"individual": True,
		# 	"enc_in" : 7
		# }

		 # Initialize actor and critic networks
		self.actor = policy_class(self.obs_dim, self.act_dim, hidden_size=self.hidden_size, hidden_size_2=self.hidden_size_2, num_layers=self.num_layers, batch_size=self.timesteps_per_batch, output_scale_factor=self.output_scale_factor )                                                   # ALG STEP 1
		self.critic = critic_class(self.obs_dim, self.act_dim, hidden_size=self.hidden_size, hidden_size_2=self.hidden_size_2, num_layers=self.num_layers, batch_size=self.timesteps_per_batch, output_scale_factor=self.output_scale_factor )                                            
		# self.actor = policy_class(actor_configs)
		# self.critic = critic_class(input_dim=7)
		# self.actor.to(self.device)
		# self.critic.to(self.device)


		# for name, param in self.actor.named_parameters():
		# 	print(f'{name}: {param.data.shape}')	

		# for name, param in self.actor.named_parameters():
		# 	print(f'{name}: {param.data}')	
		# 	print('************************************')
		
		# # he initialization for actor and critic network 
		# layer1_node = self.actor.layer1.out_features
		# layer3_node = self.actor.layer3.out_features
		# layer4_node = self.actor.layer4.out_features
		# layer5_node = self.actor.layer5.out_features
		# layer6_node = self.actor.layer6.out_features
		# layer7_node = self.actor.layer7.out_features
		# layer8_node = self.critic.layer8.out_features
		# layer9_node = self.critic.layer9.out_features
		# layer10_node = self.critic.layer10.out_features

		# layer7_node = self.actor.layer7.out_features
		# # layer8_node = self.actor.layer8.out_features
		# layer_dims_vector_actor = [self.hidden_size * 2, layer3_node, layer4_node, layer5_node, layer6_node, layer7_node, self.act_dim]
		# # layer_dims_vector_critic = [self.hidden_size * 2]
		# params_dist2 = self.initialize_parameters_he(layer_dims_vector_actor)
		# # params_dist3 = self.initialize_parameters_he(layer_dims_vector_critic)
		# with torch.no_grad():
		# # 	# self.actor.layer1.weight = nn.Parameter(torch.tensor(params_dist2['Weight1'], dtype=torch.float))
		# 	self.actor.layer3.weight = nn.Parameter(torch.tensor(params_dist2['Weight1'], dtype=torch.float))
		# 	self.actor.layer4.weight = nn.Parameter(torch.tensor(params_dist2['Weight2'], dtype=torch.float))
		# 	self.actor.layer5.weight = nn.Parameter(torch.tensor(params_dist2['Weight3'], dtype=torch.float))
		# 	self.actor.layer6.weight = nn.Parameter(torch.tensor(params_dist2['Weight4'], dtype=torch.float))
		# 	self.actor.layer7.weight = nn.Parameter(torch.tensor(params_dist2['Weight5'], dtype=torch.float))
		# # 	self.actor.layer8.weight = nn.Parameter(torch.tensor(params_dist2['Weight6'], dtype=torch.float))

		# # 	# self.actor.layer1.bias = nn.Parameter(torch.tensor(params_dist2['bias1'], dtype=torch.float).squeeze())
		# 	self.actor.layer3.bias = nn.Parameter(torch.tensor(params_dist2['bias1'], dtype=torch.float).squeeze())
		# 	self.actor.layer4.bias = nn.Parameter(torch.tensor(params_dist2['bias2'], dtype=torch.float).squeeze())
		# 	self.actor.layer5.bias = nn.Parameter(torch.tensor(params_dist2['bias3'], dtype=torch.float).squeeze())
		# 	self.actor.layer6.bias = nn.Parameter(torch.tensor(params_dist2['bias4'], dtype=torch.float).squeeze())
		# 	self.actor.layer7.bias = nn.Parameter(torch.tensor(params_dist2['bias5'], dtype=torch.float).squeeze())
		# 	self.actor.layer8.bias = nn.Parameter(torch.tensor(params_dist2['bias6'], dtype=torch.float))

			# self.critic.layer3.weight = nn.Parameter(torch.tensor(params_dist3['Weight1'], dtype=torch.float))
			# self.critic.layer4.weight = nn.Parameter(torch.tensor(params_dist3['Weight2'], dtype=torch.float))
			# self.critic.layer5.weight = nn.Parameter(torch.tensor(params_dist3['Weight3'], dtype=torch.float))
			# self.critic.layer6.weight = nn.Parameter(torch.tensor(params_dist3['Weight4'], dtype=torch.float))
			# self.critic.layer7.weight = nn.Parameter(torch.tensor(params_dist3['Weight5'], dtype=torch.float))
			# self.critic.layer8.weight = nn.Parameter(torch.tensor(params_dist3['Weight6'], dtype=torch.float))
			# self.critic.layer9.weight = nn.Parameter(torch.tensor(params_dist3['Weight7'], dtype=torch.float))
			# self.critic.layer10.weight = nn.Parameter(torch.tensor(params_dist3['Weight8'], dtype=torch.float))

			# self.critic.layer3.bias = nn.Parameter(torch.tensor(params_dist3['bias1'], dtype=torch.float).squeeze())
			# self.critic.layer4.bias = nn.Parameter(torch.tensor(params_dist3['bias2'], dtype=torch.float).squeeze())
			# self.critic.layer5.bias = nn.Parameter(torch.tensor(params_dist3['bias3'], dtype=torch.float).squeeze())
			# self.critic.layer6.bias = nn.Parameter(torch.tensor(params_dist3['bias4'], dtype=torch.float).squeeze())
			# self.critic.layer7.bias = nn.Parameter(torch.tensor(params_dist3['bias5'], dtype=torch.float).squeeze())
			# self.critic.layer8.bias = nn.Parameter(torch.tensor(params_dist3['bias6'], dtype=torch.float).squeeze())
			# self.critic.layer9.bias = nn.Parameter(torch.tensor(params_dist3['bias7'], dtype=torch.float).squeeze())
			# self.critic.layer10.bias = nn.Parameter(torch.tensor(params_dist3['bias8'], dtype=torch.float))

		self.actor.to(self.device)
		self.critic.to(self.device)

		# xavier_normal initialization 
		self.critic.apply(self.initialize_parameters_xavier)
		self.actor.apply(self.initialize_parameters_xavier)
		# for name, param in self.critic.named_parameters():
		# 	if 'weight' in name and 'layer' in name:
		# 		nn.init.xavier_normal_(param)  
		# 	elif 'bias' in name and 'layer' in name:
		# 		nn.init.constant_(param, 0)  

		# for name, param in self.actor.named_parameters():
		# 	if 'weight' in name and 'layer' in name:
		# 		nn.init.xavier_normal_(param)  
		# 	elif 'bias' in name and 'layer' in name:
		# 		nn.init.constant_(param, 0)  


		# self.critic.apply(self.initialize_parameters_xavier)

		# orthogonal initialization 
		# nn.init.orthogonal_(self.actor.layer1.weight)
		# nn.init.orthogonal_(self.actor.layer3.weight)
		# nn.init.orthogonal_(self.actor.layer4.weight)
		# nn.init.orthogonal_(self.actor.layer5.weight)
		# nn.init.orthogonal_(self.actor.layer6.weight)
		# nn.init.orthogonal_(self.actor.layer7.weight)
		# nn.init.orthogonal_(self.actor.layer8.weight)
		# nn.init.constant_(self.actor.layer1.weight, 10)
		# nn.init.constant_(self.actor.layer3.weight, 1)
		# nn.init.constant_(self.actor.layer4.weight, 1)
		# nn.init.constant_(self.actor.layer5.weight, 1)
		# nn.init.constant_(self.actor.layer6.weight, 1)
		# nn.init.constant_(self.actor.layer7.weight, 1)
		# nn.init.constant_(self.actor.layer8.weight, 1)
		# nn.init.constant_(self.actor.layer1.bias, 0)
		nn.init.constant_(self.actor.layer3.bias, 0)
		nn.init.constant_(self.actor.layer4.bias, 0)
		nn.init.constant_(self.actor.layer5.bias, 0)
		nn.init.constant_(self.actor.layer6.bias, 0)
		nn.init.constant_(self.actor.layer7.bias, 0)
		nn.init.constant_(self.actor.layer8.bias, 0)
		# nn.init.constant_(self.actor.layer9.bias, 0.5)
		# nn.init.constant_(self.actor.layer10.bias, 0.5)


		# nn.init.orthogonal_(self.critic.layer1.weight)
		# nn.init.orthogonal_(self.critic.layer2.weight)
		# nn.init.orthogonal_(self.critic.layer3.weight)
		# nn.init.orthogonal_(self.critic.layer4.weight)
		# nn.init.orthogonal_(self.critic.layer5.weight)
		# nn.init.orthogonal_(self.critic.layer6.weight)
		# nn.init.constant_(self.critic.layer1.bias, 0)
		# nn.init.constant_(self.critic.layer2.bias, 0)
		nn.init.constant_(self.critic.layer3.bias, 0)
		nn.init.constant_(self.critic.layer4.bias, 0)
		nn.init.constant_(self.critic.layer5.bias, 0)
		nn.init.constant_(self.critic.layer6.bias, 0)
		nn.init.constant_(self.critic.layer7.bias, 0)
		# nn.init.constant_(self.critic.layer8.bias, 0)
		# nn.init.constant_(self.critic.layer9.bias, 0)
		# nn.init.constant_(self.critic.layer10.bias, 0)


		# self.cov_var = torch.full(size=(self.act_dim,), fill_value=10.0, device=self.device)
		# self.cov_mat = torch.diag(self.cov_var).to(device=self.device)

		# for layer in range(1,2):
		# 	getattr(self.actor, 'weight').data = torch.tensor(params_dist['Weight'+str(layer)])
		# 	self.actor[layer].bias.data = torch.tensor(params_dist['bias'+str(layer+1)])

		# Initialize optimizers for actor and critic
		self.actor_optim = Adam(self.actor.parameters() ,lr=self.lr)
		# self.critic_optim = Adam(self.critic.parameters() ,lr=1e-3  ,betas=(0.9, 0.999))
		self.critic_optim = Adam(self.critic.parameters() ,lr=3e-4)

		# learning rate decay 
		# self.scheduler_actor = torch.optim.lr_scheduler.ReduceLROnPlateau(self.actor_optim, mode='max', factor=0.5, patience=15, verbose=True, threshold=3, threshold_mode='abs')

		# data preprocessing robustscaler 
		# self.robustscaler = RobustScaler()

		# tensorboard info 
		self.write = tensorboard_write
		self.total_step = 0

		#excel info 
		self.excel_sheet = excel_sheet

		# This logger will help us with printing out summaries of each iteration
		self.logger = {
			'delta_t': time.time_ns(),
			't_so_far': 0,          # timesteps so far
			'i_so_far': 0,          # iterations so far
			'batch_lens': [],       # episodic lengths in batch
			'batch_rews': [],       # episodic returns in batch
			'non_scaling_batch_rews': [],
			'actor_losses': [],     # losses of actor network in current iteration
			'critic_losses': [],
			'STD_ut': []
		}

	def learn(self, total_timesteps):
		"""
			Parameters:
			total_timesteps - the total number of timesteps to train for
		"""
		
		print(f"Learning... Running {self.max_timesteps_per_episode} timesteps per episode, ", end='')
		print(f"{self.timesteps_per_batch} timesteps per batch for a total of {total_timesteps} timesteps")
		t_so_far = 0 # Timesteps simulated so far
		i_so_far = 0 # Iterations ran so far
		while t_so_far < total_timesteps:                                                                       # ALG STEP 2

			# for name, param in self.actor.named_parameters():
			# 	print(f'{name}: {param.data}')	

			# print('**********************************/n')

			# for name, param in self.critic.named_parameters():
			# 	print(f'{name}: {param.data}')	

			# Initialize the covariance matrix used to query the actor for actions
			if i_so_far >= 400 :
				self.cov_var = torch.full(size=(self.act_dim,), fill_value=10.0, device=self.device)
			elif i_so_far >= 350 :
				self.cov_var = torch.full(size=(self.act_dim,), fill_value=50.0, device=self.device)
			elif i_so_far >= 300 :
				self.cov_var = torch.full(size=(self.act_dim,), fill_value=100.0, device=self.device)
			elif i_so_far >= 250 :
				self.cov_var = torch.full(size=(self.act_dim,), fill_value=350.0, device=self.device)
			elif i_so_far >= 200 :
				self.cov_var = torch.full(size=(self.act_dim,), fill_value=500.0, device=self.device)
			elif i_so_far >= 140 :
				self.cov_var = torch.full(size=(self.act_dim,), fill_value=650.0, device=self.device)
			else : 
				self.cov_var = torch.full(size=(self.act_dim,), fill_value=700.0, device=self.device)
			self.cov_mat = torch.diag(self.cov_var).to(device=self.device)

			batch_obs, batch_acts, batch_log_probs, batch_lens, batch_rews,  batch_masks, first_hidden, second_hidden = self.rollout()                     # ALG STEP 3

			# first_hidden = torch.stack(first_hidden, dim=1).squeeze()
			# second_hidden = torch.stack(second_hidden, dim=1).squeeze()

			# Calculate how many timesteps we collected this batch
			t_so_far += np.sum(batch_lens)

			# Increment the number of iterations
			i_so_far += 1

			# Logging timesteps so far and iterations so far
			self.logger['t_so_far'] = t_so_far
			self.logger['i_so_far'] = i_so_far

			# Calculate advantage at k-th iteration
			V, v_prime ,_ = self.evaluate(batch_obs, batch_acts, first_hidden, second_hidden)
			A_k, delta_v= self.compute_gae(V, v_prime, batch_masks, batch_rews, self.gamma, self.gae_lmbda)					# ALG STEP 5
			# A_k_orginal = A_k

			# batch_rtgs = self.compute_rtgs(batch_episode_rews) 
			# print('BEFORE =', A_k)                     

			# Normalizing advantages
			# but in practice it decreases the variance of our advantages and makes convergence much more stable and faster.
			# print('mean = ', A_k.mean())
			# print('STD = ', A_k.std())
			A_k = (A_k - A_k.mean()) / (A_k.std() + 1e-10)
	

			# Robust scalr data processing 
			# A_k_cpu = A_k.cpu()
			# A_k_numpy = A_k_cpu.numpy()
			# A_k_r = self.robustscaler.fit_transform(A_k_numpy.reshape(-1, 1)).flatten()
			# A_k_numpy = torch.tensor(A_k_numpy, device=self.device)

			# A_k_normalization = A_k.cpu().detach().clone()
			# A_k_normalization = A_k_normalization.numpy()
			# A_k_numpy = A_k_orginal.cpu().detach().clone()
			# A_k_numpy = A_k_numpy.numpy()
			# V_cpu = V.cpu().detach().clone()
			# delta_v_cpu = delta_v.cpu().detach().clone()
			# V_cpu = V_cpu.numpy()
			# delta_v_cpu = delta_v_cpu.numpy()
			# i_so_far_str = str(i_so_far)

			# file_path = "/home/linjim/ppo/train_data/compare_debug.xlsx"
			# expriment_name = 'sensitive analysis (two AC) -' + i_so_far_str
			# wb = load_workbook(file_path)
			# wb.create_sheet(expriment_name)
			# sheet = wb[expriment_name]
			# sheet.cell(1,1).value = 'A_K'
			# sheet.cell(1,2).value = 'A_k_normalization'
			# sheet.cell(1,3).value = 'detla_v'
			# sheet.cell(1,4).value = 'V'

			# j = 1
			# h = 1
			# y = 1
			# p = 1 
			# for i in A_k_numpy: 
			# 	j = j + 1
			# 	i = i.item()
			# 	sheet.cell(j,1).value = i
			# for k in A_k_normalization: 
			# 	h = h + 1 
			# 	k = k.item()
			# 	sheet.cell(h,2).value = k
			# for f in delta_v_cpu: 
			# 	y = y + 1 
			# 	f = f.item()
			# 	sheet.cell(y,3).value = f
			# for l in V_cpu: 
			# 	p = p + 1 
			# 	l = l.item()
			# 	sheet.cell(p,4).value = l

			# wb.save(file_path)


			# df = pd.DataFrame(A_k_r)
			# df.to_excel('A_K_r.xlsx', index=False)

			# min_range = -10
			# max_range = 10
			# desired_precision = 10
			# A_k_normalized = ((A_k_r - np.min(A_k_r)) / (np.max(A_k_r) - np.min(A_k_r))) * (max_range - min_range) + min_range
			# A_k_normalized_extended = A_k_normalized * 1e5
			# A_k_normalized_rounded = np.around(A_k_normalized_extended, decimals=desired_precision)
			# A_k = A_k_normalized_rounded / 1e5

			# A_k = torch.tensor(A_k, device=self.device)

			# A_k =  F.normalize(A_k, dim=0)
			# A_k = (A_k - A_k.min()) / (A_k.max() - A_k.min())
			# print('After =', A_k)  

			batch_log_probs = torch.stack(batch_log_probs, dim=0)

			torch.cuda.empty_cache()

			# This is the loop where we update our network for some n epochs
			for j in range(self.n_updates_per_iteration):                                                     # ALG STEP 6 & 7
				
				for i in range(40):

					# Calculate V_phi and pi_theta(a_t | s_t)
					V, _, curr_log_probs = self.evaluate(batch_obs, batch_acts, first_hidden, second_hidden)
					critic_loss = nn.MSELoss()(V, delta_v)

					print(f'critic_loss[ {i} ] = {critic_loss} ')

					# Calculate gradients and perform backward propagation for critic network
					self.critic_optim.zero_grad()
					critic_loss.backward(retain_graph=True)
					# if self.use_grad_clip: 
					# 	torch.nn.utils.clip_grad_norm_(self.critic.parameters(), 100)
					self.critic_optim.step()

					self.logger['critic_losses'].append(critic_loss.detach().to(device='cpu'))


					# for name, param in self.critic.named_parameters():
					# 	if param.grad is not None:
					# 		gradient_norm = torch.norm(param.grad)
					# 		print(f'critic Layer: {name}, Gradient Norm: {gradient_norm}')
				# Calculate V_phi and pi_theta(a_t | s_t)
					
					torch.cuda.empty_cache()


				# Calculate the ratio pi_theta(a_t | s_t) / pi_theta_k(a_t | s_t)
				ratios = torch.exp(curr_log_probs - batch_log_probs)
				# Calculate surrogate losses.
				surr1 = ratios * A_k
				surr2 = torch.clamp(ratios, 1 - self.clip, 1 + self.clip) * A_k

				# Calculate actor and critic losses.
				actor_loss = (-torch.min(surr1, surr2)).mean()
				# critic_loss = nn.MSELoss()(V, delta_v)


				# Calculate gradients and perform backward propagation for actor network
				self.actor_optim.zero_grad()
				actor_loss.backward(retain_graph=True)

				# gradient clip 
				if self.use_grad_clip: 
					torch.nn.utils.clip_grad_norm_(self.actor.parameters(), 2)
				self.actor_optim.step()

				for name, param in self.actor.named_parameters():
					if param.grad is not None:
						gradient_norm = torch.norm(param.grad)
						print(f'actor Layer: {name}, Gradient Norm: {gradient_norm}')


				# # Calculate gradients and perform backward propagation for critic network
				# self.critic_optim.zero_grad()
				# critic_loss.backward()
				# self.critic_optim.step()

				# Log actor loss

				self.logger['actor_losses'].append(actor_loss.detach().to(device='cpu'))

				torch.cuda.empty_cache()
			
			#learning rate decay
			ave_rew = torch.mean(batch_rews)
			# self.scheduler_actor.step(ave_rew)
			new_actor_lr = self.actor_optim.param_groups[0]['lr']
			self.write.add_scalar('actor_lr', float(new_actor_lr), i_so_far)
				

			# Print a summary of our training so far
			self._log_summary()

			# Save our model if it's time
			if i_so_far % self.save_freq == 0:
				torch.save(self.actor.state_dict(), './ppo_actor.pth')
				torch.save(self.critic.state_dict(), './ppo_critic.pth')

			torch.cuda.empty_cache()


	def rollout(self):
		"""tensorboard --logdir=
            on-ploicy current , need to change to off-policy 
			Parameters:
				None

			Return:
				batch_obs - the observations collected this batch. Shape: (number of timesteps, dimension of observation)
				batch_acts - the actions collected this batch. Shape: (number of timesteps, dimension of action)
				batch_log_probs - the log probabilities of each action taken this batch. Shape: (number of timesteps)
				batch_rtgs - the Rewards-To-Go of each timestep in this batch. Shape: (number of timesteps)
				batch_lens - the lengths of each episode this batch. Shape: (number of episodes)
		"""
		# Batch data. For more details, check function header.
		h_in_batch =[]
		h_out_batch =[]
		batch_obs = []
		batch_acts = []
		batch_log_probs = []
		batch_rews = []
		np_batch_rews = []
		non_scaling_rew_batch = []
		batch_rtgs = []
		batch_acts_np = []
		batch_lens = []
		batch_masks = []
		np_non_scaling_rew_batch = []
		masks = []

		np.random.seed()
		random.seed()

		# self.Reward_scaling = RewardScaling(gamma=0.95, shape=1)

		# Episodic data.  Keeps track of rewards per episode, will get cleared
		# upon each new episode

		t = 0 # Keeps track of how many timesteps we've run so far this batch

		# let reward_scaling warm mean,std 
		# while reward_scaling_warm == 0 and t < self.timesteps_per_batch : 
		# 	ep_rews = [] # rewards collected per episode
		# 	non_scaling_rew = []
		# 	masks = []
			

		# 	# adjust phi(autocorrlation)
		# 	phi_1 = round(np.random.uniform(low=0.2, high=0.9), 1)
		# 	phi_2 = round(np.random.uniform(low=0.2, high=0.9), 1)

		# 	while phi_1 <= phi_2 :
		# 		phi_1 = round(np.random.uniform(low=0.2, high=0.9), 1)
		# 		phi_2 = round(np.random.uniform(low=0.2, high=0.9), 1)
		# 		if phi_1 > phi_2 :
		# 			break

		# 	# Reset the environment. sNote that obs is short for observation. 
		# 	obs = self.env.reset(phi_1, phi_2, self.max_timesteps_per_episode)
		# 	done = False

		# 	# (num_directions * num_layers, batch_size, hidden_size) hidden_size
		# 	h_out = (torch.zeros([self.num_layers, 1, self.hidden_size_2], dtype=torch.float, device=self.device), 
		# 				torch.zeros([self.num_layers, 1, self.hidden_size_2], dtype=torch.float, device=self.device))
		# 	# h_out = torch.zeros([self.num_layers, 1, self.hidden_size_2], dtype=torch.float, device=self.device).detach()

		# 	# normalization obs & reward_scaling initialization 
		# 	# State_Normalization = Normalization(shape=(1,self.obs_dim))
		# 	self.Reward_scaling.reset()

			# for ep_t in np.arange(1, self.max_timesteps_per_episode + 1):

			# 	h_in = h_out

			# 	t += 1 # Increment timesteps ran this batch so far (include episode timestep)

			# 	# convert obs size to fit the neural network (seq_len, batch_size, input_size)
			# 	obs = obs.reshape(15, 1, self.obs_dim)
			# 	# obs = np.transpose(obs)
			
			# 	# Normalization obs 
			# 	# obs = State_Normalization.__call__(obs)
			# 	obs = torch.from_numpy(obs).to(device=self.device, dtype=torch.float32)

			# 	# Calculate action and make a step in the env. 
			# 	# rew in this episode 
			# 	action, log_prob, h_out = self.get_action(obs, h_in)
			# 	obs, rew, done, yt_target, _ = self.env.step(action, phi_1, phi_2, ep_t, self.max_timesteps_per_episode)    
			# 	mask = done
			# 	mask = not mask

			# 	# reward scaling 
			# 	scaling_rew = self.Reward_scaling.__call__(rew)

		t = 0 


		# Keep simulating until we've run more than or equal to specified timesteps per batch
		while t < self.timesteps_per_batch:
			ep_rews = [] # rewards collected per episode
			non_scaling_rew = []
			masks = []
			

			# adjust phi(autocorrlation)
			phi_1 = round(np.random.uniform(low=0.2, high=0.9), 1)
			phi_2 = round(np.random.uniform(low=0.2, high=0.9), 1)

			while phi_1 <= phi_2 :
				phi_1 = round(np.random.uniform(low=0.2, high=0.9), 1)
				phi_2 = round(np.random.uniform(low=0.2, high=0.9), 1)


			# Reset the environment. sNote that obs is short for observation. 
			obs = self.env.reset(phi_1, phi_2, self.max_timesteps_per_episode)
			done = False

			# (num_directions * num_layers, batch_size, hidden_size) hidden_size
			h_out = (torch.zeros([self.num_layers * 2, 1, self.hidden_size], dtype=torch.float, device=self.device), 
						torch.zeros([self.num_layers * 2, 1, self.hidden_size], dtype=torch.float, device=self.device))
			# h_out = torch.zeros([self.num_layers, 1, self.hidden_size_2], dtype=torch.float, device=self.device).detach()

			# normalization obs & reward_scaling initialization 
			# self.Reward_scaling.reset()


			# Run an episode for a maximum of max_timesteps_per_episode timesteps
			for ep_t in np.arange(1, self.max_timesteps_per_episode + 1):

				h_in = h_out

				t += 1 # Increment timesteps ran this batch so far (include episode timestep)

				# convert obs size to fit the neural network (seq_len, batch_size, input_size)
				obs = obs.reshape(1, 12, self.obs_dim)   	# batch first = true
				# obs = np.transpose(obs)
			
				# Normalization obs 
				# obs = State_Normalization.__call__(obs)
				obs = torch.from_numpy(obs).to(device=self.device, dtype=torch.float32)

				# Track observations in this batch
				batch_obs.append(obs)

				# Calculate action and make a step in the env. 
				# rew in this episode 
				action, log_prob, h_out = self.get_action(obs, h_in)
				obs, rew, done, yt_target, _ = self.env.step(action, phi_1, phi_2, ep_t, self.max_timesteps_per_episode)    
				mask = done
				mask = not mask

				# reward scaling 
				# scaling_rew = self.Reward_scaling.__call__(rew)


				# Track recent reward, action, and action log probability
				ep_rews.append(torch.tensor(rew))
				non_scaling_rew.append(rew)
				batch_acts_np.append(action)
				batch_acts.append(torch.tensor(action))
				batch_log_probs.append(log_prob.mean())
				masks.append(mask)
				h_in_batch.append(h_in)
				h_out_batch.append(h_out)

					# tensorboard 
				self.total_step += 1
				self.write.add_scalar('ut', float(action), self.total_step)
				self.write.add_scalar('rew', float(rew), self.total_step)
					# self.write.add_scalar('scaling rew', float(scaling_rew), self.total_step)
				self.write.add_scalar('phi_1', float(phi_1), self.total_step)
				self.write.add_scalar('phi_2', float(phi_2), self.total_step)
				self.write.add_scalar('yt_target', float(yt_target), self.total_step)
				self.write.flush()

				# If the environment tells us the episode is terminated, break
				if done == True : 
					break

			# Track episodic lengths and rewards
			batch_lens.append(ep_t)

			# batch_masks.append(masks)

			for i in ep_rews:
				batch_rews.append(i)
				np_batch_rews.append(i.item())
			for i in non_scaling_rew:
				non_scaling_rew_batch.append(i)
				np_non_scaling_rew_batch.append(i.item())
			for i in masks :
				batch_masks.append(i*1)

		std_ut = np.std(batch_acts_np)

		# Reshape data as tensors in the shape specified in function description, before returning
		batch_obs = [i.detach().clone().to(device=self.device) for i in batch_obs]
		# batch_acts = [torch.tensor(i, dtype=torch.float, device=self.device) for i in batch_acts]
		batch_acts = [i.detach().clone().to(device=self.device) for i in batch_acts]
		batch_log_probs = [i.detach().clone().to(device=self.device) for i in batch_log_probs]
		batch_rews = [i.detach().clone().to(device=self.device).view(1) for i in batch_rews]
		# batch_rews = torch.stack(batch_rews).view(self.timesteps_per_batch)
		# batch_rews = [torch.tensor(i, dtype=torch.float) for i in batch_rews]
		batch_rews = torch.stack(batch_rews).view(self.timesteps_per_batch)
		# non_scaling_rew_batch = [torch.tensor(i, dtype=torch.float) for i in non_scaling_rew_batch]
		# non_scaling_rew_batch = torch.stack(non_scaling_rew_batch).view(self.timesteps_per_batch)

		# Log the episodic returns and episodic lengths in this batch.
		self.logger['batch_rews'] = np_batch_rews
		self.logger['non_scaling_batch_rews'] = np_non_scaling_rew_batch
		self.logger['batch_lens'] = batch_lens
		self.logger['STD_ut'] = std_ut

		torch.cuda.empty_cache()

		return batch_obs, batch_acts, batch_log_probs, batch_lens, batch_rews, batch_masks, h_in_batch, h_out_batch

	def compute_rtgs(self, batch_episode_rews):
		"""
			Compute the Reward-To-Go of each timestep in a batch given the rewards. 
			State-value-function
			Use MC control 
		"""
		# The rewards-to-go (rtg) per episode per batch to return.
		# The shape will be (num timesteps per episode)
		batch_rtgs = []

		# Iterate through each episode
		for ep_rews in reversed(batch_episode_rews):

			discounted_reward = 0 # The discounted reward so far

			# Iterate through all rewards in the episode. We go backwards for smoother calculation of each
			# discounted return (think about why it would be harder starting from the beginning)
			# insert to 0 position 
			for rew in reversed(ep_rews):
				discounted_reward = rew + discounted_reward * self.gamma
				batch_rtgs.insert(0, discounted_reward)

		# Convert the rewards-to-go into a tensor
		batch_rtgs = torch.tensor(batch_rtgs, dtype=torch.float)

		return batch_rtgs
	
	def compute_gae (self, v, v_prime, batch_masks, batch_rewards, gamma, lmbda): 
		"""
			GAE 
			input : 
				v = state_value_function 
				masks = envriment done 
				rewards = 
				gamma = discount factor
				lmbda = smoothing factor
		"""
		batch_gae = []
		batch_delta_v = []
		gae = 0
		batch_rewards = batch_rewards.clone().detach()
		batch_masks = torch.tensor(batch_masks)

		for i in reversed(range(len(batch_rewards))):
			if i == (len(batch_rewards) - 1):
				v_0 = torch.zeros(1).to(device=self.device)
				v = torch.cat((v, v_0), dim=0)

			delta = batch_rewards[i] + gamma * v_prime[i] * batch_masks[i] - v[i]
			gae = delta + gamma * lmbda * batch_masks[i] * gae
			batch_gae.insert(0, gae + v[i])
			delta_v = delta + v[i]
			batch_delta_v.insert(0, delta_v)
		
		batch_gae = torch.tensor(batch_gae, dtype=torch.float, device=self.device)
		batch_delta_v = torch.tensor(batch_delta_v, dtype=torch.float, device=self.device)

		torch.cuda.empty_cache()
		
		return batch_gae, batch_delta_v


	def get_action(self, obs, h_in):
		"""
			Queries an action from the actor network, should be called from rollout.
		"""

		# Query the actor network for a mean action
		mean, lstm_hidden = self.actor(obs, h_in)
		# mean = self.actor(obs)
		mean = torch.tensor(mean.item(), device=self.device).view(1)

		normal_dist = dist.MultivariateNormal(mean, self.cov_mat)
		# normal_dist = dist.Normal(mean, var)

		# Sample an action from the distribution
		action = normal_dist.sample()

		# Calculate the log probability for that action
		log_prob = normal_dist.log_prob(action)

		action = action.item()
		# Return the sampled action and the log probability of that action in our distribution
		return action, log_prob.detach(), lstm_hidden
	
	


	def evaluate(self, batch_obs, batch_acts, first_hidden, second_hidden):
		"""
			Estimate the values of each observation, and the log probs of
			each action in the most recent batch with the most recent
			iteration of the actor network. Should be called from learn.
		"""

		batch_obs = torch.stack(batch_obs, dim=1).squeeze()				# convert np shape to tensor 
		batch_acts = torch.stack(batch_acts, dim=0)					# same above 

		# reshape batch_obs to (num_timestep_pre_batch, batch_obs_shape[0]*[1])
		# the size that to fit with batch_rtg 
		# batch_obs_v = batch_obs.view(batch_obs.size(0), -1)
		batch_acts_v = batch_acts.view(batch_acts.size(0), -1)

		# Query critic network for a value V for each batch_obs. Shape of V should be same as batch_rtgs
		# v_prime is t + 1 state, v is t state
		V_prime, _ = self.critic(batch_obs, second_hidden)
		torch.cuda.empty_cache()
		V, _ = self.critic(batch_obs, first_hidden)
		torch.cuda.empty_cache()
		# Value function normalization 
		# V = (V - V.mean()) / (V.std() + 1e-10)
		# V_prime = (V_prime - V_prime.mean()) / (V_prime.std() + 1e-10)

		# Calculate the log probabilities of batch actions using most recent actor network.
		# This segment of code is similar to that in get_action()
		mean, _ = self.actor.mult_forward(batch_obs, first_hidden)
		mean = mean.view(batch_obs.size(0), -1)
		# mean_check = []
		# for i in mean :
		# 	if math.isnan(i): 
		# 		print("NaN exist in mean")
		# 		i = torch.zeros((1,1), dtype=torch.float) 
		# 	mean_check.append(i)
		# mean_check = torch.tensor(mean_check, dtype=torch.float, requires_grad=True).unsqueeze(1)

		normal_dist = dist.MultivariateNormal(mean, self.cov_mat)
		# normal_dist = dist.Normal(mean)
		log_probs = normal_dist.log_prob(batch_acts_v)	
		# log_probs = normal_dist.log_prob(batch_acts_v).squeeze()		

		# Return the value vector V of each observation in the batch
		# and log probabilities log_probs of each action in the batch
		torch.cuda.empty_cache()

		return V, V_prime, log_probs

	def initialize_parameters_he(self, layer_dims) : 

		"""
    	initialize weights and biad as default value.
    	Arguments:
    	layer_dimes = a list contains nodes# in each layer including input features.
    	Returns:
    	parameters = a dictionary contains weights and bias
    	"""

		np.random.seed(1) #lock random parameters
		parameters = {}
		layers = len(layer_dims) 

		# count from first hidden layer
		for layer in range(1,layers): 
			if layer == 1:
				parameters["Weight"+str(layer)] = np.random.randn(layer_dims[layer], layer_dims[layer-1])
			else :
				parameters["Weight"+str(layer)] = np.random.randn(layer_dims[layer], layer_dims[layer-1])*np.sqrt(2/layer_dims[layer-1]) # He initialization
			parameters["bias"+str(layer)] = np.zeros((layer_dims[layer],1))
        
			assert(parameters['Weight' + str(layer)].shape == (layer_dims[layer], layer_dims[layer-1]))
			assert(parameters['bias' + str(layer)].shape == (layer_dims[layer],1))
                                              
		return parameters
	
	def initialize_parameters_xavier(self, model) :
		if isinstance(model, nn.Linear):
			torch.nn.init.xavier_uniform(model.weight)
			model.bias.data.fill_(0.01)

	def orthogonal_init(self, layer, gain=1.0):
		nn.init.orthogonal_(layer.weight, gain=gain)
		nn.init.constant_(layer.bias, 0)

	def _init_hyperparameters(self, hyperparameters):
		"""
			Initialize default and custom values for hyperparameters
		"""
		
		# Initialize default values for hyperparameters
		# Algorithm hyperparameters
		self.timesteps_per_batch = 4800                 # Number of timesteps to run per batch (total per batch)
		self.max_timesteps_per_episode = 1600           # Max number of timesteps per episode 
		self.n_updates_per_iteration = 5                # Number of times to update actor/critic per iteration
		self.lr = 0.005                                 # Learning rate of actor optimizer
		self.gamma = 0.95                               # Discount factor to be applied when calculating Rewards-To-Go
		self.clip = 0.2                                 # Recommended 0.2, helps define the threshold to clip the ratio during SGA
		self.gae_lmbda = 0.1		
		self.huber_delta = 20000			
		self.use_grad_clip = True		
		self.hidden_size = 256
		self.hidden_size_2 = 128 
		self.num_layers = 1
		self.output_scale_factor=100
		self.var_scale_factor = 30

		# Miscellaneous parameters
		self.save_freq = 1                             # How often we save in number of iterations
		self.seed = None                                # Sets the seed of our program, used for reproducibility of results

		# Change any default values to custom values for specified hyperparameters
		for param, val in hyperparameters.items():
			exec('self.' + param + ' = ' + str(val))

		# Sets the seed if specified
		if self.seed != None:
			# Check if our seed is valid first
			assert(type(self.seed) == int)

			# Set the seed 
			torch.manual_seed(self.seed)
			print(f"Successfully set seed to {self.seed}")

	def _log_summary(self):
		"""
			Print to stdout what we've logged so far in the most recent batch.
		"""
		# Calculate logging values. I use a few python shortcuts to calculate each value
		# without explaining since it's not too important to PPO; feel free to look it over,
		delta_t = self.logger['delta_t']
		self.logger['delta_t'] = time.time_ns()
		delta_t = (self.logger['delta_t'] - delta_t) / 1e9
		delta_t = str(round(delta_t, 2))

		t_so_far = self.logger['t_so_far']
		i_so_far = self.logger['i_so_far']
		avg_ep_lens = np.mean(self.logger['batch_lens'])
		avg_ep_rews = np.mean(self.logger['batch_rews'])
		non_scaling_avg_ep_rew = np.mean(self.logger['non_scaling_batch_rews'])
		avg_actor_loss = np.mean(self.logger['actor_losses'])
		avg_critic_loss = np.mean(self.logger['critic_losses'])
		STD_ut = self.logger['STD_ut']

		# tensorboard loss 
		self.write.add_scalar('actor_Loss', float(avg_actor_loss), i_so_far)
		self.write.add_scalar('critic_Loss', float(avg_critic_loss), i_so_far)
		self.write.add_scalar('average episodic reward', float(avg_ep_rews), i_so_far)
		self.write.add_scalar('Non scaling average episodic reward', float(non_scaling_avg_ep_rew), i_so_far)
		self.write.add_scalar('STD_ut', float(STD_ut), i_so_far)

		# execel write 
		self.excel_sheet.cell(i_so_far + 1,1).value = avg_ep_rews 

		# Round decimal places for more aesthetic logging messages
		avg_ep_lens = str(round(avg_ep_lens, 2))
		avg_ep_rews = str(round(avg_ep_rews, 10))
		avg_actor_loss = str(round(avg_actor_loss, 10))

		#    logging statements
		print(flush=True)
		print(f"-------------------- Iteration #{i_so_far} --------------------", flush=True)
		print(f"Average Episodic Length: {avg_ep_lens}", flush=True)
		print(f"Average Episodic reward : {avg_ep_rews}", flush=True)
		print(f"Average Episodic nonscaling reward : {non_scaling_avg_ep_rew}", flush=True)
		print(f"Average Loss: {avg_actor_loss}", flush=True)
		print(f"STD_ut : {STD_ut}")
		print(f"Timesteps So Far: {t_so_far}", flush=True)
		print(f"Iteration took: {delta_t} secs", flush=True)
		print(f"------------------------------------------------------", flush=True)
		print(flush=True)

		# Reset batch-specific logging data
		self.logger['batch_lens'] = []
		self.logger['batch_rews'] = []
		self.logger['actor_losses'] = []
		self.logger['critic_losses'] = []