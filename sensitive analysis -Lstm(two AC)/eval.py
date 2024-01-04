import torch
import numpy as np 
import matplotlib.pyplot as plt 
import random
from tensorboardX import SummaryWriter
from torch.utils.tensorboard import SummaryWriter
import shutil
from normalization import Normalization
from openpyxl import load_workbook


def _log_summary(ep_len, ep_ret, asd, aad, mad, std_yt, std_ut, phi_1, phi_2 ):
		"""
			Print to stdout what we've logged so far in the most recent episode.

			Parameters:
				None

			Return:
				None
		"""

		# Round decimal places for more aesthetic logging messages
		ep_average_ret = np.mean(ep_ret)
		ep_len = str(round(ep_len, 2))
		ep_ret = str(round(ep_average_ret, 2))
		std_yt = str(round(std_yt, 4))

		# Print logging statements
		print(flush=True)
		print(f"-------------------- test  # --------------------", flush=True)
		print(f"Length: {ep_len}", flush=True)
		print(f"Reward : {ep_average_ret}", flush=True)
		print(f"phi_1 : {phi_1}", flush=True)
		print(f"phi_2 : {phi_2}", flush=True)
		print(f"ASD : {asd}", flush=True)
		print(f"AAD : {aad}", flush=True)
		print(f"MAD : {mad}", flush=True)
		print(f"STD_yt : {std_yt}", flush=True)
		print(f"STD_ut : {std_ut}", flush=True)
		print(f"------------------------------------------------------", flush=True)
		print(flush=True)

def _log_total_summary(ep_len, ep_ret, asd, asd_std, aad, mad, std_yt, std_ut, std_ut_std):
		"""
			Print to stdout what we've logged so far in the most recent episode.

			Parameters:
				None

			Return:
				None
		"""

		# Round decimal places for more aesthetic logging messages
		ep_ret = torch.tensor(ep_ret)
		ep_average_ret = torch.mean(ep_ret).item()
		ep_len = str(round(ep_len, 2))
		ep_ret = str(round(ep_average_ret, 2))
		std_yt = str(round(std_yt, 4))


		# Print logging statements
		print(flush=True)
		print(f"-------------------- test  # --------------------", flush=True)
		print(f"Length: {ep_len}", flush=True)
		print(f"Reward : {ep_average_ret}", flush=True)
		print(f"ASD_mean : {asd}", flush=True)
		print(f"ASD_std : {asd_std}", flush=True)
		print(f"AAD : {aad}", flush=True)
		print(f"MAD : {mad}", flush=True)
		print(f"STD_yt : {std_yt}", flush=True)
		print(f"STD_ut_mean : {std_ut}", flush=True)
		print(f"STD_ut_std : {std_ut_std}", flush=True)
		print(f"------------------------------------------------------", flush=True)
		print(flush=True)


def rollout(policy, env, obs_dim, max_timestep_per_episode, manytest, current_test_num, num_layers, hidden_size):
	"""
		Returns a generator to roll out each episode given a trained policy and
		environment to test on. 

	"""
	# tf_borad = SummaryWriter('./tests')
	device = torch.device("cuda:2" if torch.cuda.is_available() else "cpu")
	policy.to(device)

	np.random.seed()
	random.seed()

	# Rollout until user kills process
	while True:

		# adjust phi(autocorrlation)
		phi_1 = round(np.random.uniform(low=0.2, high=0.9), 1)
		phi_2 = round(np.random.uniform(low=0.2, high=0.9), 1)

		while phi_1 <= phi_2 :
			phi_1 = round(np.random.uniform(low=0.2, high=0.9), 1)
			phi_2 = round(np.random.uniform(low=0.2, high=0.9), 1)

		# phi = 0.3
		obs = env.reset(phi_1, phi_2, max_timestep_per_episode)
		done = False
		asd = 0 
		aad = 0
		mad_array = []

		# number of timesteps so far
		total_t = 0
		calculate_t = 0

		# Logging data
		ep_len = 0            # episodic length
		ep_ret = []            # episodic return
		yt_array = []			
		ut_array = []
		std_yt_array = []
		std_ut_array = []
		t_array = []
		total_t_array = []

		# obs normalization 
		# State_Normalization = Normalization(shape=(1, 9))

		h_out = (torch.zeros([num_layers * 2, 1, hidden_size], dtype=torch.float, device=device), 
						torch.zeros([num_layers * 2, 1, hidden_size], dtype=torch.float, device=device))

		while not done:

			total_t += 1 

			h_in = h_out

			# Query deterministic action from policy and run it
			obs = obs.reshape(1, 12, obs_dim)
			obs = torch.from_numpy(obs).to(device=device, dtype=torch.float32)

			# obs = State_Normalization.__call__(obs)
			
			action, h_out = policy(obs, h_in)
			action = action.item()
			obs, rew, done, yt_target, yt = env.step(action, phi_1, phi_2, total_t, max_timestep_per_episode)

			# tensorboard 
			# tf_borad.add_scalar('ut', float(action), t)
			# tf_borad.add_scalar('obs', float(obs[2][0]), t)
			# tf_borad.add_scalar('rew', float(rew), t)
			# tf_borad.add_scalar('yt_target', float(yt_target), t)

			# Sum all episodic rewards as we go along
			rew = torch.tensor(rew)
			ep_ret.append(rew.item())

			if total_t > 10 : 
				asd += np.square(yt_target)		#ASD 
				aad += yt_target				#AAD
				mad_array.append(yt_target)		#MAD (median)

				#variance of yt, ut 
				std_yt_array.append(yt)
				std_ut_array.append(action)
				calculate_t += 1
				t_array.append(calculate_t)

			yt_array.append(yt)
			ut_array.append(action)
			total_t_array.append(total_t)

			
		# Track episodic length
		ep_len = calculate_t
		
		asd = np.sqrt(asd / calculate_t)
		asd = asd.item()
		aad = np.sqrt(aad / calculate_t)
		aad = aad.item()
		mad = np.median(np.abs(mad_array - np.mean(mad_array)))
		std_yt = np.std(std_yt_array)
		std_ut = np.std(std_ut_array)

		# dummy = torch.randn(1,9)
		# tf_borad.add_graph(policy, dummy)
		# tf_borad.close()
		# tf_borad.flush()

		# write in excel 
		file_path = "/home/linjim/ppo/result_data/data.xlsx"
		current_test_num = str(current_test_num)
		expriment_name = 'sensitive analysis -' + current_test_num
		wb = load_workbook(file_path)
		wb.create_sheet(expriment_name)
		sheet = wb[expriment_name]
		sheet.cell(1,1).value = 'yt'
		sheet.cell(1,2).value = 'ut'

		j = 1
		h = 1
		for i in yt_array: 
			j = j + 1
			sheet.cell(j,1).value = i
		for k in ut_array: 
			h = h + 1 
			sheet.cell(h,2).value = k

		
		sheet.cell(len(yt_array)+1, 1).value = 'ASD'
		sheet.cell(len(yt_array)+2, 1).value = 'STD_yt'
		sheet.cell(len(yt_array)+3, 1).value = 'STD_ut'
		sheet.cell(len(yt_array)+1, 2).value = asd
		sheet.cell(len(yt_array)+2, 2).value = std_yt
		sheet.cell(len(yt_array)+3, 2).value = std_ut

		wb.save(file_path)


		# print graph 
		if manytest == False : 
			fig = plt.figure('graph',figsize=(7,3))
			fig.suptitle('')
			# ut_flaten = [item[0][0] for item in ut_array]
			target_array = [10 for i in range(total_t)]

			plt.subplot(2, 1, 1)
			plt.plot(total_t_array ,yt_array ,color='seagreen', label='yt')
			plt.plot(total_t_array ,target_array ,color='orangered' ,linewidth= 2 ,linestyle=':' , label='Target')
			plt.legend(['yt','target'])
			plt.xlabel('timesteps')
			plt.subplot(2, 1, 2)
			plt.plot(total_t_array ,ut_array ,color='seagreen', label='ut')
			plt.legend(['ut'])
			plt.xlabel('timesteps')
			plt.show()

		# returns episodic length and return in this iteration
		return ep_len, ep_ret, asd, aad, mad, std_yt, std_ut, phi_1, phi_2

def eval_policy(policy, env, obs_dim, max_timestep_per_episode, manytest, current_test_num, num_layers, hidden_size):

	# Rollout with the policy and environment, and log each episode's data
	shutil.rmtree('tests', ignore_errors=True)
	ep_len, ep_ret, asd, aad, mad, std_yt, std_ut, phi_1, phi_2 = rollout(policy, env, obs_dim, max_timestep_per_episode, manytest, current_test_num, num_layers, hidden_size)
	_log_summary(ep_len=ep_len, ep_ret=ep_ret, asd= asd, aad= aad, mad= mad, std_yt= std_yt, std_ut= std_ut, phi_1=phi_1, phi_2=phi_2)

	return asd, aad, mad, std_yt, std_ut