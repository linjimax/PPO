import sys
import torch
import os
import numpy as np 
import torch.nn as nn

from envriment import envri
from arguments import get_args
from ppo import PPO
from network import Actor_FeedForwardNN,Critic_FeedForwardNN
from Dlinear_network import Actor_FeedForwardNN_Dlinear_Model, Critic_FeedForwardNN_Dlinear_Model
from eval import eval_policy, _log_total_summary
from tensorboardX import SummaryWriter
from openpyxl import load_workbook

def train(env, hyperparameters, actor_model, critic_model):
	"""
		Trains the model.

		Parameters:
			env - the environment to train on
			hyperparameters - a dict of hyperparameters to use, defined in main
			actor_model - the actor model to load in if we want to continue training
			critic_model - the critic model to load in if we want to continue training

		Return:
			None
	"""	
	print(f"Training", flush=True)

	log_dir = '/home/linjim/ppo/implement/sensitive analysis -Lstm (two AC) -35'
	write = SummaryWriter(log_dir)

	# excel write
	excel_write = 'sensitive analysis -Lstm (two AC) -35'

	file_path = "/home/linjim/ppo/train_data/data(two AC).xlsx"
	wb = load_workbook(file_path)
	wb.create_sheet(excel_write)
	sheet = wb[excel_write]
	sheet.cell(1,1).value = 'average_per_iteration_reward'


	# Create a model for PPO.
	model = PPO(policy_class=Actor_FeedForwardNN, critic_class=Critic_FeedForwardNN, env=env, tensorboard_write=write, excel_sheet=sheet, **hyperparameters)


	# Tries to load in an existing actor/critic model to continue training on
	if actor_model != '' and critic_model != '' :
		print(f"Loading in {actor_model} and {critic_model}...", flush=True)
		model.actor.load_state_dict(torch.load(actor_model))                # load from ppo
		model.critic.load_state_dict(torch.load(critic_model))
		print(f"Successfully loaded.", flush=True)
	elif actor_model != '' or critic_model != '' : # Don't train from scratch if user accidentally forgets actor/critic model
		print(f"Error: Either specify both actor/critic models or none at all. We don't want to accidentally override anything!")
		sys.exit(0)
	else:
		print(f"Training from scratch.", flush=True)

 
	model.learn(total_timesteps=4500000)

	wb.save(file_path)


def test(env, hyperparameters, actor_model, manytest, current_test_num):
	"""
		Tests the model.

		Parameters:
			env - the environment to test the policy on
			actor_model - the actor model to load in

		Return:
			None
	"""
	print(f"Testing {actor_model}", flush=True)

	# If the actor model is not specified, then exit
	if os.path.getsize(actor_model) == 0 :
		print(f"Didn't specify model file. Exiting.", flush=True)
		sys.exit(0)

	# Extract out dimensions of observation and action spaces
	obs_dim = 7                                # depend on s
	act_dim = 1                                # controller ut
	m_t_p_e = hyperparameters['max_timesteps_per_episode']
	hidden_size = hyperparameters['hidden_size']
	hidden_size_2 = hyperparameters['hidden_size_2']
	num_layers = hyperparameters['num_layers']
	o_p_s_f = hyperparameters['output_scale_factor'] 

	# Build our policy the same way we build our actor model in PPO
	policy = Actor_FeedForwardNN(obs_dim, act_dim, hidden_size, hidden_size_2, num_layers, batch_size=1, output_scale_factor=o_p_s_f)

	# temp 
	# layer3_node = policy.layer3.out_features
	# layer4_node = policy.layer4.out_features
	# layer5_node = policy.layer5.out_features
	# layer6_node = policy.layer6.out_features
	# layer7_node = policy.layer7.out_features
	# layer_dims_vector_actor = [layer3_node, layer4_node, layer5_node, layer6_node, layer7_node,  act_dim]
	# params_dist2 = initialize_parameters_he(layer_dims_vector_actor)
	# policy.layer3.weight = nn.Parameter(torch.tensor(params_dist2['Weight1'], dtype=torch.float))
	# policy.layer4.weight = nn.Parameter(torch.tensor(params_dist2['Weight2'], dtype=torch.float))
	# policy.layer5.weight = nn.Parameter(torch.tensor(params_dist2['Weight3'], dtype=torch.float))
	# policy.layer6.weight = nn.Parameter(torch.tensor(params_dist2['Weight4'], dtype=torch.float))
	# policy.layer7.weight = nn.Parameter(torch.tensor(params_dist2['Weight5'], dtype=torch.float))
	# policy.layer3.bias = nn.Parameter(torch.tensor(params_dist2['bias1'], dtype=torch.float).squeeze())
	# policy.layer4.bias = nn.Parameter(torch.tensor(params_dist2['bias2'], dtype=torch.float).squeeze())
	# policy.layer5.bias = nn.Parameter(torch.tensor(params_dist2['bias3'], dtype=torch.float).squeeze())
	# policy.layer6.bias = nn.Parameter(torch.tensor(params_dist2['bias4'], dtype=torch.float).squeeze())
	# policy.layer7.bias = nn.Parameter(torch.tensor(params_dist2['bias5'], dtype=torch.float).squeeze())	

	# Load in the actor model saved by the PPO algorithm
	policy.load_state_dict(torch.load(actor_model))

	m_t_p_e = hyperparameters['max_timesteps_per_episode']
	
	asd, aad, mad, std_yt, std_ut = eval_policy(
												policy=policy, 
												env=env, 
												obs_dim = obs_dim, 
												max_timestep_per_episode = m_t_p_e, 
												manytest=manytest, 
												current_test_num=current_test_num, 
												num_layers=num_layers, 
												hidden_size=hidden_size
												)


	return asd, aad, mad, std_yt, std_ut
def main(args):
	"""
		The main function to run.

		Parameters:
			args - the arguments parsed from command line

			
		Return: 
			None
	"""
	hyperparameters = {
				'timesteps_per_batch': 10000, 
				'max_timesteps_per_episode': 200, 
				'gamma': 0.97, 
				'n_updates_per_iteration':7,
				'lr': 3e-4,
				'clip': 0.25,
				'render': True,
				'env_target': 10,
				'env_reward_epsilon':5,
				'env_reward_constant':1000,
				'gae_lmbda':0.9,
				'use_grad_clip':True,
				'hidden_size':128,
				'hidden_size_2':256,
				'num_layers':3,
				'output_scale_factor':100,
				'var_scale_factor':30
			  }


	env_target = hyperparameters['env_target']
	env_reward_epsilon = hyperparameters['env_reward_epsilon']
	env_reward_constant = hyperparameters['env_reward_constant']
	env = envri(env_target, env_reward_epsilon, env_reward_constant)

	args.actor_model = './ppo_actor.pth'
	args.critic_model = './ppo_critic.pth'

	# Train or test, depending on the mode specified
	if args.mode == 'train':
		train(env=env, hyperparameters=hyperparameters, actor_model=args.actor_model, critic_model=args.critic_model)
	elif args.mode == 'retrain' :
		args.actor_model = ''
		args.critic_model =''
		train(env=env, hyperparameters=hyperparameters, actor_model=args.actor_model, critic_model=args.critic_model)
	elif args.mode == 'manytest':
		test_unm =  50			# number of sample test 
		asd_array = []
		aad_array = []
		mad_array = []
		std_ut_array = []
		std_yt_array = []
		for i in range(test_unm): 
			asd, aad, mad, std_yt, std_ut = test(env=env, hyperparameters=hyperparameters, actor_model=args.actor_model, manytest=True, current_test_num=i)
			asd_array.append(asd)
			aad_array.append(aad)
			mad_array.append(mad)
			std_ut_array.append(std_ut)
			std_yt_array.append(std_yt)
		sample_asd = np.mean(asd_array)
		sample_asd_std = np.std(asd_array)
		sample_aad = np.mean(aad_array)
		sample_mad = np.mean(mad_array)
		sample_std_ut = np.mean(std_ut_array)
		sample_std_ut_std= np.std(std_ut_array)
		sample_std_yt = np.mean(std_yt_array)
		_log_total_summary(ep_len=float(0), ep_ret=float(0), asd=sample_asd, asd_std=sample_asd_std, aad=sample_aad, mad=sample_mad, std_yt=sample_std_yt, std_ut=sample_std_ut, std_ut_std=sample_std_ut_std)
	else : 
		test(env=env, hyperparameters=hyperparameters, actor_model=args.actor_model, manytest=False, current_test_num=6)

def initialize_parameters_he(layer_dims) : 

		"""
    	initialize weights and biad as default value.
    	Arguments:
    	layer_dimes = a list contains nodes# in each layer including input features.
    	Returns:
    	parameters = a dictionary contains weights and bias
    	"""

		np.random.seed(1) #lock random parameters
		parameters = {}
		layers = len(layer_dims) 

		# count from first hidden layer
		for layer in range(1,layers): 
			if layer == 1:
				parameters["Weight"+str(layer)] = np.random.randn(layer_dims[layer], layer_dims[layer-1])
			else :
				parameters["Weight"+str(layer)] = np.random.randn(layer_dims[layer], layer_dims[layer-1])*np.sqrt(2/layer_dims[layer-1]) # He initialization
			parameters["bias"+str(layer)] = np.zeros((layer_dims[layer],1))
        
			assert(parameters['Weight' + str(layer)].shape == (layer_dims[layer], layer_dims[layer-1]))
			assert(parameters['bias' + str(layer)].shape == (layer_dims[layer],1))
                                              
		return parameters

if __name__ == '__main__':
	args = get_args() # Parse arguments from command line
	main(args)
